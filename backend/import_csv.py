"""SENTRY Backend — CSV / Excel importer.

Reads a CSV or .xlsx file and upserts rows into the vendors table.

Expected columns (case-insensitive, flexible matching):
  company_name | Company Name     — REQUIRED
  category                        — e.g. "Drones & C-UAS"
  technology_product               — e.g. "DroneShield RFOne"
  report_url                       — link to assessment report
  overall_rating                   — numeric 0–5
  vendor_status                    — e.g. Active, Under Review
  risk_level                       — Low / Medium / High / Critical
  last_assessed                    — date string e.g. "2025-03-15"
  company_url                      — vendor website

Any missing columns get sensible defaults.
"""
import csv
import hashlib
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]

from database import get_connection, init_db

# ── Column name normalization map ─────────────────────────────────
ALIASES: dict[str, str] = {
    "company name":       "company_name",
    "company":            "company_name",
    "vendor":             "company_name",
    "vendor name":        "company_name",
    "name":               "company_name",
    "company url":        "company_url",
    "website":            "company_url",
    "url":                "company_url",
    "category":           "category",
    "technology product": "technology_product",
    "product":            "technology_product",
    "technology":         "technology_product",
    "tech product":       "technology_product",
    "report url":         "report_url",
    "report link":        "report_url",
    "report":             "report_url",
    "overall rating":     "overall_rating",
    "rating":             "overall_rating",
    "score":              "overall_rating",
    "vendor status":      "vendor_status",
    "status":             "vendor_status",
    "risk level":         "risk_level",
    "risk":               "risk_level",
    "last assessed":      "last_assessed",
    "last audited":       "last_assessed",
    "date":               "last_assessed",
    "assessed":           "last_assessed",
}

DB_COLUMNS = [
    "company_name", "company_url", "category", "technology_product",
    "report_url", "overall_rating", "vendor_status", "risk_level",
    "last_assessed",
]


def _normalize_header(raw: str) -> str:
    """Map a raw column header to our canonical DB column name."""
    key = raw.strip().lower().replace("_", " ")
    return ALIASES.get(key, key.replace(" ", "_"))


def _make_id(company_name: str, product: str) -> str:
    """Deterministic ID from company + product."""
    slug = f"{company_name}::{product}".lower().strip()
    return hashlib.sha256(slug.encode()).hexdigest()[:12]


def _parse_rating(val: str | float | None) -> float:
    """Coerce a rating value to float 0–5."""
    if val is None:
        return 0.0
    try:
        r = float(val)
        return max(0.0, min(5.0, r))
    except (ValueError, TypeError):
        return 0.0


def _normalize_risk(val: str | None) -> str:
    """Normalize risk level to one of Low/Medium/High/Critical."""
    if not val:
        return "Medium"
    v = val.strip().lower()
    mapping = {"low": "Low", "medium": "Medium", "med": "Medium",
               "high": "High", "critical": "Critical", "crit": "Critical"}
    return mapping.get(v, "Medium")


def _read_csv(path: Path) -> list[dict[str, str]]:
    """Read a CSV file and return list of row dicts."""
    rows: list[dict[str, str]] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return rows
        col_map = {col: _normalize_header(col) for col in reader.fieldnames}
        for row in reader:
            normalized = {col_map[k]: v for k, v in row.items() if k in col_map}
            rows.append(normalized)
    return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    """Read an Excel .xlsx file and return list of row dicts."""
    if load_workbook is None:
        print("ERROR: openpyxl not installed. Run: uv pip install openpyxl")
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    raw_rows = list(ws.iter_rows(values_only=True))
    if len(raw_rows) < 2:
        return []
    headers = [str(h or "").strip() for h in raw_rows[0]]
    col_map = {i: _normalize_header(h) for i, h in enumerate(headers) if h}
    rows: list[dict[str, str]] = []
    for row in raw_rows[1:]:
        d = {col_map[i]: str(cell or "") for i, cell in enumerate(row) if i in col_map}
        rows.append(d)
    wb.close()
    return rows


def import_file(path: Path) -> int:
    """Import a CSV or XLSX file into SQLite. Returns row count."""
    init_db()
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(path)
    elif suffix in (".xlsx", ".xls"):
        rows = _read_xlsx(path)
    else:
        print(f"ERROR: Unsupported file type '{suffix}'. Use .csv or .xlsx")
        sys.exit(1)

    if not rows:
        print("WARNING: No data rows found in file.")
        return 0

    conn = get_connection()
    count = 0
    for row in rows:
        name = row.get("company_name", "").strip()
        if not name:
            continue
        product = row.get("technology_product", "").strip()
        vid = _make_id(name, product)
        conn.execute(
            """INSERT OR REPLACE INTO vendors
               (id, company_name, company_url, category, technology_product,
                report_url, overall_rating, vendor_status, risk_level, last_assessed)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                vid,
                name,
                row.get("company_url", "").strip(),
                row.get("category", "Other").strip() or "Other",
                product,
                row.get("report_url", "").strip(),
                _parse_rating(row.get("overall_rating")),
                row.get("vendor_status", "Active").strip() or "Active",
                _normalize_risk(row.get("risk_level")),
                row.get("last_assessed", "").strip(),
            ),
        )
        count += 1
    conn.commit()
    conn.close()
    print(f"\u2705 Imported {count} vendors into SQLite.")
    return count


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_csv.py <vendors.csv|vendors.xlsx>")
        print("\nPlace your file in backend/data/ and run:")
        print("  python import_csv.py data/vendors.csv")
        sys.exit(1)
    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)
    import_file(filepath)
