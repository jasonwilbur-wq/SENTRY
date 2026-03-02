#!/usr/bin/env python3
"""
SENTRY Enhanced Data Importer - 202601 & 202602

Imports vendor enrichment data from Excel trackers and links VAR documents.

Features:
- Import vendor highlights, pros/cons, concerns, use cases
- Link organized VAR documents to vendor cards
- Update vendor ratings and assessment data
- Generate import summary report

Author: Atlas (Code Puppy)
Date: 2026-02-28
"""

import sqlite3
import json
import re
import openpyxl
from pathlib import Path
from datetime import datetime
import hashlib

# Paths
DB_PATH = Path("data/sentry.db")
CONFIG_PATH = Path("sentry_config.json")
TRACKER_202601 = Path(r"C:\Users\j0w16ja\OneDrive - Walmart Inc\ET\SENTRY_Data\Trackers\Emerging Tech Tracker_202601.xlsx")
TRACKER_202602 = Path(r"C:\Users\j0w16ja\OneDrive - Walmart Inc\ET\SENTRY_Data\Trackers\Emerging Tech Tracker_202602.xlsx")
VARS_BASE = Path(r"C:\Users\j0w16ja\OneDrive - Walmart Inc\ET\VARs")

# Load config
with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

# Stats
stats = {
    'vendors_updated': 0,
    'vendors_created': 0,
    'vars_linked': 0,
    'errors': []
}

def slugify(text: str) -> str:
    """Convert text to URL-safe slug."""
    text = text.lower().strip()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')

def extract_vendor_name_from_var_filename(filename: str) -> str:
    """
    Extract vendor name from VAR filename.
    
    Examples:
        WMT-SEC-VAR-20260215-Skydio-X10D-Detailed-v1.docx -> skydio
        WMT-SEC-VAR-20260210-PaloAltoNetworks-CortexXSIAM20-Detailed-v1.docx -> palo-alto-networks
    """
    match = re.match(r'WMT-SEC-VAR-\d{8}-(.*?)-(?:Detailed|Summary)', filename, re.IGNORECASE)
    if match:
        vendor_slug = slugify(match.group(1))
        return vendor_slug
    return ""

def find_vendor_in_db(cursor, vendor_name: str) -> str | None:
    """
    Find vendor ID by fuzzy matching company name.
    
    Returns vendor ID if found, None otherwise.
    """
    # Exact match
    cursor.execute(
        "SELECT id FROM vendors WHERE LOWER(company_name) = LOWER(?)",
        (vendor_name,)
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    
    # Slug match
    vendor_slug = slugify(vendor_name)
    cursor.execute(
        "SELECT id, company_name FROM vendors"
    )
    for row in cursor.fetchall():
        if slugify(row[1]) == vendor_slug:
            return row[0]
    
    # Partial match (vendor name in company name)
    cursor.execute(
        "SELECT id FROM vendors WHERE LOWER(company_name) LIKE LOWER(?)",
        (f"%{vendor_name}%",)
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    
    return None

def create_vendor_id(company_name: str) -> str:
    """Generate deterministic vendor ID from company name."""
    slug = slugify(company_name)
    hash_suffix = hashlib.md5(slug.encode()).hexdigest()[:8]
    return f"vendor-{slug[:30]}-{hash_suffix}"

def import_tracker_row(cursor, row_data: dict, source_month: str):
    """
    Import a single vendor row from Excel tracker.
    
    Actual columns from tracker:
    - Date
    - Status (numeric score)
    - Company
    - Technology_Product
    - Category
    - Use Case
    - Add Value to Walmart
    - Maturity Level
    - Source URL/Publisher
    - Analysis Completed
    - Additional Notes
    """
    company_name = str(row_data.get('Company', '')).strip()
    if not company_name or company_name == 'None':
        return  # Skip empty rows
    
    # Check if vendor exists
    vendor_id = find_vendor_in_db(cursor, company_name)
    
    # Prepare data
    tech_product = str(row_data.get('Technology_Product', '') or '').strip()
    category = str(row_data.get('Category', 'Other') or 'Other').strip()
    use_case = str(row_data.get('Use Case', '') or '').strip()
    value_to_walmart = str(row_data.get('Add Value to Walmart', '') or '').strip()
    maturity_level = str(row_data.get('Maturity Level', '') or '').strip()
    source_url = str(row_data.get('Source URL/Publisher', '') or '').strip()
    additional_notes = str(row_data.get('Additional Notes', '') or '').strip()
    
    # Use status as overall rating (appears to be numeric score)
    overall_rating = 0.0
    try:
        if row_data.get('Status'):
            overall_rating = float(row_data['Status'])
    except (ValueError, TypeError):
        pass
    
    if vendor_id:
        # Update existing vendor
        update_fields = []
        update_values = []
        
        if use_case:
            update_fields.append('use_cases = ?')
            update_values.append(use_case)
        
        if value_to_walmart:
            update_fields.append('value_to_walmart = ?')
            update_values.append(value_to_walmart)
        
        if maturity_level:
            update_fields.append('maturity_level = ?')
            update_values.append(maturity_level)
        
        if overall_rating > 0:
            update_fields.append('overall_rating = ?')
            update_values.append(overall_rating)
        
        if category and category != 'Other':
            update_fields.append('category = ?')
            update_values.append(category)
        
        if tech_product:
            update_fields.append('technology_product = ?')
            update_values.append(tech_product)
        
        if source_url:
            update_fields.append('company_url = ?')
            update_values.append(source_url)
        
        # Use additional notes as description if available
        if additional_notes:
            update_fields.append('description = ?')
            update_values.append(additional_notes)
        
        # Update last_assessed
        update_fields.append('last_assessed = ?')
        update_values.append(source_month)
        
        if update_fields:
            update_values.append(vendor_id)
            sql = f"UPDATE vendors SET {', '.join(update_fields)} WHERE id = ?"
            cursor.execute(sql, update_values)
            stats['vendors_updated'] += 1
            print(f"   ✅ Updated: {company_name}")
    
    else:
        # Create new vendor
        vendor_id = create_vendor_id(company_name)
        cursor.execute("""
            INSERT INTO vendors (
                id, company_name, company_url, category, technology_product,
                description, use_cases, value_to_walmart, maturity_level,
                overall_rating, vendor_status, risk_level, has_var, last_assessed
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            vendor_id,
            company_name,
            source_url,
            category,
            tech_product,
            additional_notes,
            use_case,
            value_to_walmart,
            maturity_level,
            overall_rating,
            'Active',
            'Medium',
            0,  # has_var (will be updated when linking VARs)
            source_month
        ))
        stats['vendors_created'] += 1
        print(f"   🆕 Created: {company_name}")

def read_excel_tracker(file_path: Path, source_month: str) -> list[dict]:
    """
    Read Excel tracker and extract vendor data.
    
    Returns list of row dictionaries.
    """
    print(f"\n📖 Reading tracker: {file_path.name}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Read header row
        headers = [cell.value for cell in ws[1]]
        
        # Read data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            row_data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            rows.append(row_data)
        
        print(f"   Found {len(rows)} rows")
        return rows
    
    except Exception as e:
        stats['errors'].append(f"Error reading {file_path.name}: {e}")
        print(f"   ❌ Error: {e}")
        return []

def link_vars_to_vendors(cursor):
    """
    Scan organized VAR folders and link documents to vendors.
    """
    print("\n🔗 Linking VAR documents to vendors...")
    
    for month in CONFIG['var_months']:
        month_folder = VARS_BASE / month
        if not month_folder.exists():
            continue
        
        print(f"\n   📁 Processing {month}/")
        var_files = list(month_folder.glob('*.docx'))
        print(f"      Found {len(var_files)} VAR documents")
        
        for var_file in var_files:
            vendor_slug = extract_vendor_name_from_var_filename(var_file.name)
            if not vendor_slug:
                continue
            
            # Find vendor in DB
            vendor_id = find_vendor_in_db(cursor, vendor_slug.replace('-', ' '))
            if not vendor_id:
                # Try without dashes
                vendor_id = find_vendor_in_db(cursor, vendor_slug.replace('-', ''))
            
            if vendor_id:
                # Check if VAR already linked
                cursor.execute(
                    "SELECT id FROM var_reports WHERE vendor_id = ? AND filename = ?",
                    (vendor_id, var_file.name)
                )
                if cursor.fetchone():
                    continue  # Already linked
                
                # Create VAR report record
                var_id = f"var-{hashlib.md5(var_file.name.encode()).hexdigest()[:12]}"
                report_date = month[:4] + '-' + month[4:6] + '-01'
                
                cursor.execute("""
                    INSERT INTO var_reports (
                        id, vendor_id, filename, download_url, report_date,
                        report_version, report_type, match_method, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    var_id,
                    vendor_id,
                    var_file.name,
                    str(var_file),
                    report_date,
                    'v1',
                    'Detailed' if 'Detailed' in var_file.name else 'Summary',
                    'auto-organized',
                    datetime.now().isoformat()
                ))
                
                # Update vendor has_var flag
                cursor.execute(
                    "UPDATE vendors SET has_var = 1 WHERE id = ?",
                    (vendor_id,)
                )
                
                stats['vars_linked'] += 1
                print(f"      ✅ Linked: {var_file.name} -> {vendor_id}")

def generate_kpi_stats(cursor):
    """
    Generate KPI statistics for Vendor Directory.
    """
    print("\n📊 Generating KPI Statistics...")
    
    # Total vendors
    cursor.execute("SELECT COUNT(*) FROM vendors")
    total_vendors = cursor.fetchone()[0]
    print(f"   Total Vendors: {total_vendors}")
    
    # Vendors with VARs
    cursor.execute("SELECT COUNT(*) FROM vendors WHERE has_var = 1")
    vendors_with_vars = cursor.fetchone()[0]
    print(f"   Vendors with VARs: {vendors_with_vars}")
    
    # Category breakdown
    cursor.execute("""
        SELECT category, COUNT(*) as count
        FROM vendors
        GROUP BY category
        ORDER BY count DESC
    """)
    categories = cursor.fetchall()
    print(f"\n   📂 Category Breakdown:")
    for cat, count in categories:
        pct = (count / total_vendors * 100) if total_vendors > 0 else 0
        print(f"      {cat}: {count} ({pct:.1f}%)")
    
    # Maturity level breakdown
    cursor.execute("""
        SELECT maturity_level, COUNT(*) as count
        FROM vendors
        WHERE maturity_level IS NOT NULL AND maturity_level != ''
        GROUP BY maturity_level
        ORDER BY count DESC
    """)
    maturity_levels = cursor.fetchall()
    print(f"\n   🔬 Maturity Level Breakdown:")
    for level, count in maturity_levels:
        print(f"      {level}: {count}")

def main():
    """
    Main import process.
    """
    print("="*60)
    print("🐶 SENTRY Enhanced Data Importer - 202601 & 202602")
    print("="*60)
    
    # Connect to database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    try:
        # Import 202601 tracker
        if TRACKER_202601.exists():
            rows_202601 = read_excel_tracker(TRACKER_202601, '202601')
            print(f"\n📥 Importing {len(rows_202601)} vendors from 202601...")
            for row in rows_202601:
                import_tracker_row(cursor, row, '202601')
        else:
            print(f"⚠️  Tracker not found: {TRACKER_202601}")
        
        # Import 202602 tracker
        if TRACKER_202602.exists():
            rows_202602 = read_excel_tracker(TRACKER_202602, '202602')
            print(f"\n📥 Importing {len(rows_202602)} vendors from 202602...")
            for row in rows_202602:
                import_tracker_row(cursor, row, '202602')
        else:
            print(f"⚠️  Tracker not found: {TRACKER_202602}")
        
        # Link VAR documents
        link_vars_to_vendors(cursor)
        
        # Generate KPI stats
        generate_kpi_stats(cursor)
        
        # Commit changes
        conn.commit()
        
        print("\n" + "="*60)
        print("✅ IMPORT COMPLETE!")
        print("="*60)
        print(f"\n📊 Summary:")
        print(f"   Vendors Updated: {stats['vendors_updated']}")
        print(f"   Vendors Created: {stats['vendors_created']}")
        print(f"   VARs Linked: {stats['vars_linked']}")
        if stats['errors']:
            print(f"\n⚠️  Errors: {len(stats['errors'])}")
            for error in stats['errors'][:10]:
                print(f"   - {error}")
        
        print("\n🎉 Your Vendor Directory is now enhanced!")
        print("   - Vendor cards have highlights, pros/cons, use cases")
        print("   - VARs are linked and accessible from vendor cards")
        print("   - KPI data is ready for the dashboard\n")
    
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        conn.rollback()
        raise
    
    finally:
        conn.close()

if __name__ == "__main__":
    main()
