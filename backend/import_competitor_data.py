#!/usr/bin/env python3
"""Import SENTRY competitor intelligence workbooks.

Discovers monthly Walmart_Competitor_YYYYMM.xlsx files from the OneDrive
SENTRY data-entry workspace, normalizes their Incident_log rows, scores each
signal, and refreshes the competitor_events / competitor_entities tables.

Safe to re-run: the import refreshes the competitor tables from discovered
source workbooks so stale partial data does not linger.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from competitor_scoring import score_event  # noqa: E402
from database import get_connection, init_db  # noqa: E402

ONEDRIVE_ROOT = Path(r"C:\Users\j0w16ja\OneDrive - Walmart Inc")
SOURCE_DIRS = [
    ONEDRIVE_ROOT / "Data Entries" / "SENTRY" / "Competitor Data",
    ONEDRIVE_ROOT / "Desktop" / "SENTRY",
    Path(__file__).parent / "data" / "source" / "competitor",
]

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

EXCLUDE_ENTITIES = {
    "", "walmart", "walmart (vicinity)", "sam's club", "sams club",
    "industry", "retail industry", "retailers (general)", "retail (general)",
    "competitor", "global (general)", "logistics (general)", "logistics sector",
    "tech sector", "general retail", "federal govt", "nist", "cisa",
}

CATEGORY_PATTERNS: list[tuple[str, str]] = [
    (r"cyber|breach|hack|malware|ransomware|phishing|credential|ddos|vulnerability", "Cyber"),
    (r"orc|organized retail crime|theft|robbery|shoplift|cargo|shrink", "ORC/Theft"),
    (r"recall|contamination|food.?safety|product safety", "Recall"),
    (r"legal|lawsuit|settlement|litigation|court", "Legal"),
    (r"regulatory|compliance|fine|violation|privacy law|gdpr|ccpa", "Compliance"),
    (r"strategic|acquisition|partnership|expansion|market|launch", "Strategic"),
    (r"operational|store.?operations|supply.?chain|outage|disruption|labor|strike", "Operational"),
    (r"technology|tech|ai|automation|robot|drone|computer vision|rfid", "Technology"),
    (r"fraud|scam|identity.?theft|payment", "Fraud"),
]

COLUMN_MAP = {
    "date": "event_date",
    "competitor/entity": "competitor",
    "competitor": "competitor",
    "entity": "competitor",
    "event title": "event_title",
    "title": "event_title",
    "event type": "event_type",
    "type": "event_type",
    "detailed description": "detailed_description",
    "description": "detailed_description",
    "category": "category",
    "location/geographic scope": "location",
    "location": "location",
    "security implication": "security_implication",
    "operational impact": "operational_impact",
    "financial impact": "financial_impact",
    "reputational impact": "reputational_impact",
    "source/link": "source_link",
    "source link": "source_link",
    "source": "source_link",
    "analyst notes": "analyst_notes",
    "confidence": "confidence_level",
    "confidence level": "confidence_level",
}

BASE_COLUMNS = [
    "event_date", "competitor", "event_title", "event_type", "detailed_description",
    "category", "location", "security_implication", "operational_impact",
    "financial_impact", "reputational_impact", "source_link", "analyst_notes", "source_month",
]

SCORE_COLUMNS = [
    "walmart_relevance_score", "priority_tier", "signal_type", "recommended_owner",
    "why_walmart_cares", "strategic_score", "security_score", "operational_score",
    "customer_trust_score", "novelty_score", "urgency_score", "confidence_score",
    "escalate_to_cso", "scoring_version", "confidence_level", "score_reason",
    "confidence_effect", "source_effect", "cso_candidate_reason",
]


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", cell_str(value).lower())


def month_label_from_path(path: Path) -> str:
    match = re.search(r"(20\d{2})(\d{2})", path.stem)
    if not match:
        return path.stem
    year = int(match.group(1))
    month = int(match.group(2))
    if 1 <= month <= 12:
        return f"{MONTH_NAMES[month - 1]} {year}"
    return path.stem


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    raw = cell_str(value)
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", raw)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return raw[:10]


def normalize_competitor(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", name).strip()
    lowered = cleaned.lower()
    if lowered in {"aws", "amazon (aws)", "amazon (corp)", "amazon (retail)", "amazon fresh", "amazon (ring)"} or lowered.startswith("amazon"):
        return "Amazon"
    if lowered in {"aldi", "aldi us", "aldi (us)"}:
        return "ALDI"
    if lowered.startswith("lidl"):
        return "Lidl"
    if "zebra" in cleaned or lowered.startswith("evri"):
        return "Zebra Technologies"
    return cleaned


def is_valid_competitor(name: str) -> bool:
    lowered = name.strip().lower()
    if lowered in EXCLUDE_ENTITIES:
        return False
    if "walmart" in lowered or "sam's club" in lowered or "sams club" in lowered:
        return False
    if any(generic in lowered for generic in ("general)", "industry", "sector", "multiple)")):
        return False
    return True


def normalize_category(category: str, event_title: str, description: str, event_type: str) -> str:
    clean = cell_str(category)
    allowed = {"Cyber", "ORC/Theft", "Recall", "Legal", "Strategic", "Operational", "Compliance", "Fraud", "Technology", "Other"}
    if clean in allowed:
        return clean
    if clean == "Regulatory":
        return "Compliance"

    combined = f"{clean} {event_title} {description} {event_type}".lower()
    for pattern, label in CATEGORY_PATTERNS:
        if re.search(pattern, combined, re.IGNORECASE):
            return label
    return "Other"


def discover_workbooks() -> list[Path]:
    candidates: dict[str, Path] = {}
    for source_dir in SOURCE_DIRS:
        if not source_dir.exists():
            continue
        for path in source_dir.glob("Walmart_Competitor_*.xlsx"):
            if path.name.startswith("~$"):
                continue
            candidates[path.stem] = path
    return [candidates[key] for key in sorted(candidates)]


def read_workbook(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Incident_log"] if "Incident_log" in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_header(h) for h in next(rows_iter)]
    except StopIteration:
        return []

    mapped_headers = [COLUMN_MAP.get(h, "") for h in headers]
    source_month = month_label_from_path(path)
    records: list[dict[str, Any]] = []

    for raw_row in rows_iter:
        if not raw_row or all(cell is None or cell_str(cell) == "" for cell in raw_row):
            continue
        record = {col: "" for col in BASE_COLUMNS}
        extras: dict[str, str] = {}
        for idx, cell in enumerate(raw_row[: len(mapped_headers)]):
            mapped = mapped_headers[idx]
            value = cell_str(cell)
            if mapped:
                record[mapped] = value
            elif headers[idx]:
                extras[headers[idx]] = value

        competitor = normalize_competitor(record.get("competitor", ""))
        if not is_valid_competitor(competitor):
            continue

        record["event_date"] = normalize_date(record.get("event_date"))
        record["competitor"] = competitor
        record["category"] = normalize_category(
            record.get("category", ""),
            record.get("event_title", ""),
            record.get("detailed_description", ""),
            record.get("event_type", ""),
        )
        record["source_month"] = source_month
        if not record.get("event_title") and record.get("detailed_description"):
            record["event_title"] = record["detailed_description"][:90]
        if not record.get("event_title"):
            continue

        confidence = record.pop("confidence_level", "") or extras.get("confidence", "") or extras.get("confidence level", "")
        scored = score_event({**record, "confidence_level": confidence})
        for col in SCORE_COLUMNS:
            record[col] = scored.get(col, confidence if col == "confidence_level" else "")
        records.append(record)

    print(f"  {path.name}: {len(records)} clean events ({source_month})")
    return records


def create_tables(conn) -> None:
    # init_db creates canonical tables; these indexes are kept for older DBs.
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ce_competitor ON competitor_events(competitor)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ce_date ON competitor_events(event_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ce_category ON competitor_events(category)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ce_source_month ON competitor_events(source_month)")
    conn.commit()


def threat_level(event_count: int, cyber_count: int) -> str:
    ratio = cyber_count / max(event_count, 1)
    if event_count >= 30 or ratio >= 0.20:
        return "High"
    if event_count >= 10 or ratio >= 0.10:
        return "Medium"
    return "Low"


def refresh_competitor_entities(conn) -> None:
    rows = conn.execute("SELECT competitor, category, source_month FROM competitor_events WHERE deleted_at IS NULL").fetchall()
    grouped: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for competitor, category, source_month in rows:
        grouped[competitor].append((category or "Other", source_month or ""))

    conn.execute("DELETE FROM competitor_entities")
    for name, items in grouped.items():
        categories = Counter(category for category, _ in items)
        monthly = Counter(month for _, month in items if month)
        total = len(items)
        top_category = categories.most_common(1)[0][0] if categories else "Other"
        conn.execute(
            """
            INSERT INTO competitor_entities
              (name, event_count, cyber_count, orc_count, recall_count, legal_count,
               strategic_count, threat_level, top_category, categories_json, monthly_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                total,
                categories.get("Cyber", 0),
                categories.get("ORC/Theft", 0),
                categories.get("Recall", 0),
                categories.get("Legal", 0),
                categories.get("Strategic", 0),
                threat_level(total, categories.get("Cyber", 0)),
                top_category,
                json.dumps(dict(categories), sort_keys=True),
                json.dumps(dict(sorted(monthly.items())), sort_keys=True),
            ),
        )
    conn.commit()
    print(f"  Refreshed {len(grouped)} competitor entities")


def main() -> None:
    print("\n" + "=" * 70)
    print("SENTRY Competitor Data Import")
    print("=" * 70)

    workbooks = discover_workbooks()
    if not workbooks:
        raise FileNotFoundError(f"No Walmart_Competitor_*.xlsx files found in: {SOURCE_DIRS}")

    print("Discovered workbooks:")
    for path in workbooks:
        print(f"  - {path}")

    events: list[dict[str, Any]] = []
    for workbook in workbooks:
        events.extend(read_workbook(workbook))

    init_db()
    conn = get_connection()
    create_tables(conn)
    conn.execute("DELETE FROM competitor_events")

    columns = BASE_COLUMNS + SCORE_COLUMNS
    placeholders = ",".join(["?"] * len(columns))
    conn.executemany(
        f"INSERT INTO competitor_events ({','.join(columns)}) VALUES ({placeholders})",
        [[event.get(col, "") for col in columns] for event in events],
    )
    conn.commit()
    print(f"\nInserted {len(events)} competitor events")

    refresh_competitor_entities(conn)

    stats = conn.execute(
        "SELECT COUNT(*), COUNT(DISTINCT competitor), MIN(event_date), MAX(event_date) FROM competitor_events"
    ).fetchone()
    print(f"Events: {stats[0]} | Competitors: {stats[1]} | Date range: {stats[2]} → {stats[3]}")
    print("Top competitors:")
    for name, count in conn.execute(
        "SELECT competitor, COUNT(*) FROM competitor_events GROUP BY competitor ORDER BY COUNT(*) DESC LIMIT 10"
    ):
        print(f"  {name}: {count}")
    conn.close()
    print("\n✅ Competitor import complete\n")


if __name__ == "__main__":
    main()
