"""
build_regulatory_report.py
==========================
Build the regulatory intelligence briefing from the live 2026 workbook when
available, otherwise fall back to the legacy 202601/202602 CLEAN files.

Outputs:
  - data/json_reports/regulatory-briefing.json
  - data/regulatory_rows.json

Run from backend/ directory.
"""
import csv
import json
import re
import hashlib
import sys
from datetime import datetime, date
from pathlib import Path

from path_config import REGULATORY_ROOT

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl',
        '--index-url', 'https://pypi.ci.artifacts.walmart.com/artifactory/api/pypi/external-pypi/simple',
        '--allow-insecure-host', 'pypi.ci.artifacts.walmart.com', '-q'])
    import openpyxl

REGULATORY_SOURCE_DIR = REGULATORY_ROOT
WORKBOOK_2026 = REGULATORY_SOURCE_DIR / 'Regulatory Data - 2026.xlsx'
WORKBOOK_COPY = REGULATORY_SOURCE_DIR / 'Regulatory Data - Copy.xlsx'
LOCAL_WORKBOOK_2026 = Path('data/source/Regulatory Data - 2026.xlsx')

OUT_JSON   = Path('data/json_reports/regulatory-briefing.json')
OUT_RAW    = Path('data/regulatory_rows.json')
OUT_CLEAN_JSON = Path('data/source/regulatory_cleaned_master.json')
OUT_CLEAN_CSV = Path('data/source/regulatory_cleaned_master.csv')
NOW_ISO    = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')

# ── helpers ────────────────────────────────────────────────────────────────────
def slug(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')[:60]

def cell_str(v) -> str:
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def read_sheet_rows(path: Path, worksheet_name: str | None = None, prefer_clean: bool = False) -> list[dict]:
    """Read rows from a workbook sheet into dicts with provenance metadata."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if worksheet_name:
        sheet = wb[worksheet_name]
    elif prefer_clean:
        sheet = next((ws for ws in wb.worksheets if 'CLEAN' in ws.title.upper()), wb.worksheets[0])
    else:
        sheet = wb.worksheets[0]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [cell_str(c) or f'col_{i}' for i, c in enumerate(rows[0])]
    out: list[dict] = []
    for ri, row in enumerate(rows[1:], start=2):
        if all(c is None for c in row):
            continue
        item = {h: cell_str(v) for h, v in zip(headers, row)}
        item['_source'] = f'{path.name}:{sheet.title}:row{ri}'
        item['_sheet'] = sheet.title
        out.append(item)
    return out

REQUIRED_COLS = {
    'Location',
    'Type of Technology',
    'Name of Law, Regulation, Ordinance, or Bill',
    'Status',
    'Detailed Description of Scope and Key Provisions',
    'Source Link(s)',
}

DATE_ALIASES = {
    'Date': 'Date Enacted or Proposed',
    'Date Enacted or Proposed': 'Date Enacted or Proposed',
}


def _normalize_headers(row0: list) -> list[str]:
    headers: list[str] = []
    for i, raw in enumerate(row0):
        clean = cell_str(raw)
        clean = DATE_ALIASES.get(clean, clean)
        headers.append(clean or f'col_{i}')
    return headers


def _sheet_rows(path: Path, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = _normalize_headers(list(rows[0]))
    if not REQUIRED_COLS.issubset(set(headers)):
        return []

    out: list[dict] = []
    for ri, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        item = {h: cell_str(v) for h, v in zip(headers, row)}
        if not item.get('Name of Law, Regulation, Ordinance, or Bill'):
            continue
        item['_source'] = f'{path.name}:{sheet_name}:row{ri}'
        item['_sheet'] = sheet_name
        out.append(item)
    return out


def _read_workbook_all_valid_tabs(path: Path) -> tuple[list[dict], dict[str, int], str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []
    counts: dict[str, int] = {}

    for ws in wb.worksheets:
        tab_rows = _sheet_rows(path, ws.title)
        if not tab_rows:
            continue
        counts[f'{path.name}:{ws.title}'] = len(tab_rows)
        rows.extend(tab_rows)

    month_tabs = sorted(
        [ws.title for ws in wb.worksheets if re.fullmatch(r'2026\d{2}', ws.title)]
    )
    data_through = f"{month_tabs[-1][:4]}-{month_tabs[-1][4:6]}-30" if month_tabs else ''
    return rows, counts, data_through


def load_source_rows() -> tuple[list[dict], dict[str, int], str, str]:
    """Load regulatory rows from Desktop/SENTRY/Regulatory across all relevant tabs."""
    all_rows: list[dict] = []
    counts: dict[str, int] = {}
    data_through = ''

    source_files = []
    for candidate in (WORKBOOK_2026, WORKBOOK_COPY, LOCAL_WORKBOOK_2026):
        if candidate.exists():
            source_files.append(candidate)

    if not source_files:
        raise FileNotFoundError(
            f'No source workbook found in {REGULATORY_SOURCE_DIR} or backend/data/source'
        )

    for src in source_files:
        rows, local_counts, workbook_through = _read_workbook_all_valid_tabs(src)
        all_rows.extend(rows)
        counts.update(local_counts)
        if workbook_through and workbook_through > data_through:
            data_through = workbook_through

    return all_rows, counts, (data_through or '2026-12-31'), 'desktop_sentry_regulatory_workbooks'

# ── canonical jurisdiction normaliser ──────────────────────────────────────────
JURIS_MAP = {
    'united states (federal)': 'United States (Federal)',
    'united states': 'United States (Federal)',
    'usa': 'United States (Federal)',
    'us': 'United States (Federal)',
    'usa (federal)': 'United States (Federal)',
    'us federal': 'United States (Federal)',
    'federal': 'United States (Federal)',
    'federal, us': 'United States (Federal)',
    'federal,usa': 'United States (Federal)',
    'federal (us)': 'United States (Federal)',
    'federal, usa': 'United States (Federal)',
    'us (federal)': 'United States (Federal)',
    'eu': 'European Union',
    'europe (eu)': 'European Union',
    'european union': 'European Union',
    'uk': 'United Kingdom',
    'united kingdom': 'United Kingdom',
    'global': 'Global',
    'international': 'Global',
}

US_STATES = [
    'Alabama','Alaska','Arizona','Arkansas','California','Colorado','Connecticut',
    'Delaware','Florida','Georgia','Hawaii','Idaho','Illinois','Indiana','Iowa',
    'Kansas','Kentucky','Louisiana','Maine','Maryland','Massachusetts','Michigan',
    'Minnesota','Mississippi','Missouri','Montana','Nebraska','Nevada',
    'New Hampshire','New Jersey','New Mexico','New York','North Carolina',
    'North Dakota','Ohio','Oklahoma','Oregon','Pennsylvania','Rhode Island',
    'South Carolina','South Dakota','Tennessee','Texas','Utah','Vermont',
    'Virginia','Washington','West Virginia','Wisconsin','Wyoming','District Of Columbia',
]

US_STATE_ABBR = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District Of Columbia',
}

COUNTRY_MAP = {
    'australia': 'Australia', 'new zealand': 'New Zealand', 'south korea': 'South Korea',
    'korea': 'South Korea', 'china': 'China', 'hong kong': 'Hong Kong', 'india': 'India',
    'japan': 'Japan', 'philippines': 'Philippines', 'vietnam': 'Vietnam',
    'taiwan': 'Taiwan', 'germany': 'Germany', 'ireland': 'Ireland', 'spain': 'Spain',
    'canada': 'Canada', 'brazil': 'Brazil', 'nigeria': 'Nigeria', 'south africa': 'South Africa',
    'pakistan': 'Pakistan', 'algeria': 'Algeria', 'russia': 'Russia', 'montenegro': 'Montenegro',
    'turkey': 'Turkey', 'italy': 'Italy', 'scotland': 'United Kingdom', 'bangladesh': 'Bangladesh',
    'ecuador': 'Ecuador', 'singapore': 'Singapore', 'thailand': 'Thailand', 'portugal': 'Portugal',
    'caribbean': 'Global', 'navajo nation': 'United States (Federal)',
}


def _us_state_from_text(text: str) -> str | None:
    upper = text.upper()
    for abbr, state in US_STATE_ABBR.items():
        if re.search(rf'(^|[\s,;/\-]){abbr}([\s,;/\-]|$)', upper):
            return state

    for state in US_STATES:
        if state.lower() in text:
            return state

    if 'new york city' in text or text in {'nyc, us', 'new york, ny', 'newyork'} or 'poestenkill' in text:
        return 'New York'
    if any(v in text for v in ['los angeles', 'san francisco', 'san jose', 'santa cruz', 'san diego', 'oakland', 'woodland']):
        return 'California'
    if any(v in text for v in ['portland', 'eugene']):
        return 'Oregon'
    if any(v in text for v in ['cambridge', 'brookline']):
        return 'Massachusetts'
    if 'chicago' in text:
        return 'Illinois'
    if 'denver' in text:
        return 'Colorado'
    if 'minot' in text:
        return 'North Dakota'
    if 'district of columbia' in text or text == 'dc':
        return 'District Of Columbia'

    return None


def norm_juris(raw: str) -> str:
    if not raw:
        return 'Unknown'

    clean = re.sub(r'\s+', ' ', raw.strip())
    lo = clean.lower()

    if lo in JURIS_MAP:
        return JURIS_MAP[lo]

    for alias, canonical in JURIS_MAP.items():
        if alias in lo:
            return canonical

    us_state = _us_state_from_text(lo)
    if us_state:
        return f'{us_state}, USA'

    for key, country in COUNTRY_MAP.items():
        if key in lo:
            return country

    if ' / ' in clean or 'multiple states' in lo:
        return 'Global'

    if lo in {'unknown', 'bedford county'}:
        return 'Unknown (City Level)'

    return clean.title()

# ── tech category normaliser ────────────────────────────────────────────────────
TECH_PRIORITY = {
    'AI': 5, 'Biometrics': 5, 'ALPR/LPR': 4, 'Drones/UAS': 4,
    'Data Privacy': 4, 'Surveillance': 3, 'ORC': 3,
    'Weapons Detection': 3, 'Robotics': 2, 'Other': 1,
}

def norm_tech(raw: str) -> str:
    if not raw:
        return 'Other'
    lo = raw.lower()
    if 'ai' in lo or 'artificial' in lo or 'machine learning' in lo:
        return 'AI'
    if 'biometric' in lo or 'facial' in lo or 'fingerprint' in lo:
        return 'Biometrics'
    if 'alpr' in lo or 'license plate' in lo or 'lpr' in lo:
        return 'ALPR/LPR'
    if 'drone' in lo or 'uas' in lo or 'uav' in lo:
        return 'Drones/UAS'
    if 'privacy' in lo or 'data protection' in lo or 'gdpr' in lo:
        return 'Data Privacy'
    if 'surveillance' in lo or 'cctv' in lo or 'camera' in lo:
        return 'Surveillance'
    if 'orc' in lo or 'retail crime' in lo:
        return 'ORC'
    if 'weapon' in lo or 'firearm' in lo:
        return 'Weapons Detection'
    if 'robot' in lo or 'amr' in lo:
        return 'Robotics'
    return 'Other'

# ── status normaliser ──────────────────────────────────────────────────────────
STATUS_MAP = {
    'enacted': 'Enacted',
    'enforced': 'Enacted',
    'enforcement': 'Enacted',
    'active': 'Enacted',
    'in effect': 'Enacted',
    'approved': 'Enacted',
    'signed': 'Enacted',
    'proposed': 'Proposed',
    'pending': 'Proposed',
    'introduced': 'Proposed',
    'approved by committee': 'Proposed',
    'released': 'Proposed',
    'passed': 'Enacted',
    'failed': 'Failed',
    'withdrawn': 'Failed',
    'expired': 'Failed',
}

def norm_status(raw: str) -> str:
    lo = raw.lower().strip()
    for k, v in STATUS_MAP.items():
        if k in lo:
            return v
    return 'Proposed'

# ── risk scoring engine ────────────────────────────────────────────────────────
def score_risk(row: dict) -> dict:
    """Heuristic risk scoring based on tech category, status, and jurisdiction."""
    tech    = row.get('_tech', 'Other')
    status  = row.get('_status', 'Proposed')
    juris   = row.get('_jurisdiction', '')
    desc    = (row.get('Detailed Description of Scope and Key Provisions', '') or '').lower()

    # Impact: based on tech sensitivity and operational breadth
    impact_base = TECH_PRIORITY.get(tech, 2)
    # Boost if description mentions penalties, fines, enforcement
    if any(w in desc for w in ['penalty', 'fine', 'civil', 'criminal', 'enforcement', 'prohibition']):
        impact_base = min(5, impact_base + 1)
    # Federal / EU / UK = higher enterprise impact
    if any(x in juris for x in ['Federal', 'European Union', 'United Kingdom']):
        impact_base = min(5, impact_base + 1)
    impact = max(1, min(5, impact_base))

    # Likelihood: enacted laws are almost certain; proposed are possible-likely
    if status == 'Enacted':
        likelihood = 4
    elif status == 'Proposed':
        likelihood = 3
    else:
        likelihood = 1
    # Reduce for non-Walmart-relevant jurisdictions (international with no ops)
    if juris in ('European Union', 'United Kingdom') and 'United States' not in juris:
        likelihood = max(1, likelihood - 1)

    score = impact * likelihood
    if score <= 6:   rag = 'Green'
    elif score <= 12: rag = 'Yellow'
    elif score <= 18: rag = 'Amber'
    else:            rag = 'Red'

    reason = (
        f"Tech={tech}, Status={status}, Jurisdiction={juris}: "
        f"Impact={impact}/5 (operational breadth + enforcement provisions), "
        f"Likelihood={likelihood}/5 (enacted={status=='Enacted'}). "
        f"Score={score}."
    )
    return {'impact': impact, 'likelihood': likelihood, 'score': score, 'rag': rag, 'reason': reason}

# ── NIST CSF framework mapper ──────────────────────────────────────────────────
NIST_MAP = {
    'AI':               ('NIST AI RMF GOVERN-1.1', 'GS-AI-01'),
    'Biometrics':       ('NIST CSF PR.AC-1',        'GS-BIO-01'),
    'ALPR/LPR':         ('NIST CSF PR.AC-3',        'GS-ALPR-01'),
    'Drones/UAS':       ('NIST CSF PR.IP-1',        'GS-UAS-01'),
    'Data Privacy':     ('ISO 27001 A.18.1.4',       'GS-PRIV-01'),
    'Surveillance':     ('NIST CSF DE.CM-1',        'GS-CAM-01'),
    'ORC':              ('NIST CSF DE.CM-3',        'GS-ORC-01'),
    'Weapons Detection':('NIST CSF PR.PT-1',        'GS-WEAP-01'),
    'Robotics':         ('NIST CSF PR.IP-2',        'GS-ROBOT-01'),
    'Other':            ('NIST CSF ID.GV-1',        'GS-GOV-01'),
}

# ── evidence status (heuristic from status+description) ───────────────────────
def evidence_status(status: str, desc: str) -> str:
    if status == 'Enacted':
        return 'Partially'   # assumed until full VAR/compliance review
    return 'Unknown'

# ── ingest workbook / fallback files ──────────────────────────────────────────
all_rows, source_counts, data_through, source_mode = load_source_rows()
print(f'Loaded regulatory source mode: {source_mode}')
for sheet_name, count in source_counts.items():
    print(f'  {sheet_name}: {count} rows')
print(f'  Combined: {len(all_rows)} rows before dedup')

# ── deduplicate by (jurisdiction + law name + tech) with merge semantics ─────
merged_rows: dict[str, dict] = {}
for row in all_rows:
    loc = cell_str(row.get('Location', ''))
    name = cell_str(row.get('Name of Law, Regulation, Ordinance, or Bill', ''))
    tech = norm_tech(cell_str(row.get('Type of Technology', '')))
    jurisdiction = norm_juris(loc)
    status = norm_status(cell_str(row.get('Status', '')))
    key = slug(f'{jurisdiction}|{name}|{tech}')

    row['_jurisdiction'] = jurisdiction
    row['_tech'] = tech
    row['_status'] = status
    row['_id'] = 'REG-' + hashlib.md5(key.encode()).hexdigest()[:8].upper()
    row['_key'] = key
    row['_provenance'] = [row['_source']]

    existing = merged_rows.get(key)
    if not existing:
        merged_rows[key] = row
        continue

    # Prefer the richer row while preserving provenance.
    if len(json.dumps(row, default=str)) > len(json.dumps(existing, default=str)):
        survivor = row
        previous = existing
    else:
        survivor = existing
        previous = row

    for field, value in previous.items():
        if field.startswith('_'):
            continue
        if not survivor.get(field) and value:
            survivor[field] = value

    survivor['_provenance'] = sorted(set(existing.get('_provenance', []) + row.get('_provenance', [])))
    merged_rows[key] = survivor

deduped = list(merged_rows.values())
print(f'  After dedup: {len(deduped)} unique obligations')

# ── sort by risk descending ────────────────────────────────────────────────────
for row in deduped:
    row['_risk'] = score_risk(row)

deduped.sort(key=lambda r: r['_risk']['score'], reverse=True)

# ── build jurisdiction list ────────────────────────────────────────────────────
jurisdictions = sorted(set(r['_jurisdiction'] for r in deduped))

# ── build obligations list for JSON ───────────────────────────────────────────
def _geo_scope(jurisdiction: str) -> tuple[str, str | None, str | None]:
    if jurisdiction.endswith(', USA'):
        state = jurisdiction.replace(', USA', '')
        state_code = next((abbr for abbr, name in US_STATE_ABBR.items() if name == state), None)
        return 'US_STATE', state, state_code
    if jurisdiction == 'United States (Federal)':
        return 'US_FEDERAL', 'United States', None
    if jurisdiction == 'Global':
        return 'GLOBAL', None, None
    return 'COUNTRY', jurisdiction, None


def build_obligation(row: dict) -> dict:
    tech  = row['_tech']
    nist_ctrl, gs_tag = NIST_MAP.get(tech, ('NIST CSF ID.GV-1', 'GS-GOV-01'))
    desc  = row.get('Detailed Description of Scope and Key Provisions', '') or ''
    # truncate summary to ~30 words
    words = desc.split()
    summary_30 = ' '.join(words[:30]) + ('...' if len(words) > 30 else '')

    raw_link = row.get('Source Link(s)', '') or ''
    links = [l.strip() for l in re.split(r'[,;\n]+', raw_link) if l.strip().startswith('http')]
    if not links:
        links = []

    ev_status = evidence_status(row['_status'], desc.lower())

    # date parsing
    raw_date = row.get('Date Enacted or Proposed', '') or ''
    enacted_date = None
    if isinstance(raw_date, (datetime, date)):
        enacted_date = raw_date.strftime('%Y-%m-%d') if hasattr(raw_date, 'strftime') else str(raw_date)
    elif raw_date:
        # Try to parse YYYY-MM-DD
        m = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', raw_date)
        if m:
            enacted_date = f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
        else:
            # Try YYYY
            m2 = re.search(r'(\d{4})', raw_date)
            enacted_date = f'{m2.group(1)}-01-01' if m2 else None

    # criticality based on risk score
    risk_score = row['_risk']['score']
    criticality = 1 if risk_score <= 4 else 2 if risk_score <= 8 else 3 if risk_score <= 12 else 4 if risk_score <= 18 else 5

    controls = [{
        'control_id':   nist_ctrl,
        'description':  f'{tech} compliance review and evidence collection',
        'owner':        'Global Security & Emerging Technology',
        'status':       'Partial' if ev_status == 'Partially' else 'None',
        'last_reviewed': '2026-02-28',
        'evidence_link': links[0] if links else ''
    }, {
        'control_id':   gs_tag,
        'description':  f'Internal {tech} governance control',
        'owner':        'Global Security & Emerging Technology',
        'status':       'Partial',
        'last_reviewed': '2026-02-28',
        'evidence_link': links[1] if len(links) > 1 else ''
    }]

    geo_scope, state_name, state_code = _geo_scope(row['_jurisdiction'])

    return {
        'id':             row['_id'],
        'jurisdiction':   row['_jurisdiction'],
        'title':          row.get('Name of Law, Regulation, Ordinance, or Bill', 'Unknown'),
        'summary':        summary_30,
        'tech_category':  tech,
        'effective_date': enacted_date,
        'deadline':       None,
        'criticality':    criticality,
        'evidence_status': ev_status,
        'evidence_links': links,
        'risk':           row['_risk'],
        'controls':       controls,
        'full_description': desc,
        'status':         row['_status'],
        'provenance':     row.get('_provenance', [row['_source']]),
        'geo_scope':      geo_scope,
        'country':        None if geo_scope.startswith('US_') else row['_jurisdiction'],
        'state':          state_name,
        'state_code':     state_code,
    }

obligations = [build_obligation(r) for r in deduped]

# ── executive summary (top 3 actions) ─────────────────────────────────────────
top3_red_amber = [o for o in obligations if o['risk']['rag'] in ('Red', 'Amber')][:3]
exec_summary = (
    f"As of {datetime.utcnow().strftime('%B %Y')}, SENTRY's regulatory tracker covers "
    f"{len(obligations)} unique obligations across {len(jurisdictions)} jurisdictions, spanning "
    f"AI, Biometrics, ALPR/LPR, Drones/UAS, and Data Privacy technologies. "
    f"{sum(1 for o in obligations if o['risk']['rag'] == 'Red')} Red and "
    f"{sum(1 for o in obligations if o['risk']['rag'] == 'Amber')} Amber obligations require immediate attention. "
    f"Top 3 actions: (1) Validate AI governance controls against EU AI Act and state-level RAISE Act requirements; "
    f"(2) Review ALPR/LPR data retention compliance for enacted state laws; "
    f"(3) Confirm biometric consent and disclosure posture across all US store deployments."
)

# ── top actions ───────────────────────────────────────────────────────────────
top_actions = [
    {
        'title': 'AI Governance Review',
        'description': 'Conduct cross-functional review of AI use cases against EU AI Act, NY RAISE Act, and Colorado AI Act requirements. Map controls to NIST AI RMF.',
        'owner': 'Global Security & Emerging Technology / Legal',
        'priority': 'High',
        'eta': '2026-04-30'
    },
    {
        'title': 'Biometric Consent Audit',
        'description': 'Audit biometric data collection at all US store locations for BIPA (IL), TX CUBI, WA MY Health MY Data, and NYC Admin Code compliance.',
        'owner': 'Privacy / Legal / Store Ops',
        'priority': 'High',
        'eta': '2026-04-15'
    },
    {
        'title': 'ALPR Data Retention Policy Update',
        'description': 'Update ALPR data retention schedules to comply with enacted state laws (VA, MN, NM, WA). Document evidence of policy enforcement.',
        'owner': 'Global Security / Legal',
        'priority': 'High',
        'eta': '2026-05-01'
    },
    {
        'title': 'UAS / Drone Vendor Compliance Review',
        'description': 'Verify all drone vendors are not on FCC prohibited list and comply with FAA Part 107 requirements.',
        'owner': 'Global Security / Procurement',
        'priority': 'Med',
        'eta': '2026-05-15'
    },
    {
        'title': 'State-Level ORC Compliance Mapping',
        'description': 'Map CORCA (H.R. 2853) and state ORC laws to existing loss prevention programs and identify reporting obligation gaps.',
        'owner': 'Asset Protection / Legal',
        'priority': 'Med',
        'eta': '2026-06-01'
    },
]

# ── assemble JSON ──────────────────────────────────────────────────────────────
report = {
    'id':            'regulatory-briefing',
    'title':         'SENTRY Regulatory Intelligence Briefing — Global Security & Emerging Technology',
    'summary':       exec_summary,
    'created_at':    NOW_ISO,
    'data_through':  data_through,
    'jurisdictions': jurisdictions,
    'stats': {
        'total_obligations': len(obligations),
        'red':     sum(1 for o in obligations if o['risk']['rag'] == 'Red'),
        'amber':   sum(1 for o in obligations if o['risk']['rag'] == 'Amber'),
        'yellow':  sum(1 for o in obligations if o['risk']['rag'] == 'Yellow'),
        'green':   sum(1 for o in obligations if o['risk']['rag'] == 'Green'),
        'enacted': sum(1 for o in obligations if o['status'] == 'Enacted'),
        'proposed':sum(1 for o in obligations if o['status'] == 'Proposed'),
        'tech_breakdown': {}
    },
    'obligations':   obligations,
    'top_actions':   top_actions,
    'assumptions': [
        'SOURCE OF TRUTH: Use Desktop/SENTRY/Regulatory workbooks (Regulatory Data - 2026.xlsx + Regulatory Data - Copy.xlsx) and ingest all valid tabs that contain the required schema.',
        'SCHEMA MAPPING: Location -> jurisdiction, Type of Technology -> tech_category, Name of Law... -> title, Status -> evidence_status (after normalisation), Detailed Description -> summary/full_description, Date Enacted or Proposed -> effective_date, Source Link(s) -> evidence_links.',
        'No compliance_evidence.csv was provided. Control mappings are generated from NIST CSF/AI RMF and internal GS tag templates. All evidence_status values should be validated by compliance team.',
        'Deadline dates are not present in the source data; deadline fields are set to null. Legal team should populate these.',
        'Risk scores are heuristic: Impact derived from tech sensitivity + jurisdiction scope + penalty language detection. Likelihood derived from enactment status. All top-12 assessments should be validated by Legal and Security.',
        'Duplicate detection uses normalised jurisdiction + law name + technology as the composite key. Duplicate rows are merged and provenance from all source rows is retained.',
    ],
    'confidence':    'Med',
    'schema_version': '1.0',
    'ingestion_notes': {
        'source_mode': source_mode,
        **{sheet_name: f'{count} rows ingested' for sheet_name, count in source_counts.items()},
        'dedup_removed': len(all_rows) - len(deduped),
        'final_obligations': len(obligations),
    }
}

# tech breakdown stats
from collections import Counter
tech_counts = Counter(o['tech_category'] for o in obligations)
report['stats']['tech_breakdown'] = dict(tech_counts.most_common())

# ── write outputs ──────────────────────────────────────────────────────────────
OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
OUT_CLEAN_JSON.parent.mkdir(parents=True, exist_ok=True)
OUT_JSON.write_text(json.dumps(report, indent=2, default=str), encoding='utf-8')
OUT_RAW.write_text(json.dumps(deduped, indent=2, default=str), encoding='utf-8')
OUT_CLEAN_JSON.write_text(json.dumps(obligations, indent=2, default=str), encoding='utf-8')

with open(OUT_CLEAN_CSV, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=[
        'id', 'jurisdiction', 'geo_scope', 'country', 'state', 'state_code',
        'title', 'tech_category', 'status', 'effective_date',
        'risk_score', 'risk_rag', 'criticality', 'evidence_status', 'source_count',
    ])
    writer.writeheader()
    for o in obligations:
        writer.writerow({
            'id': o['id'],
            'jurisdiction': o['jurisdiction'],
            'geo_scope': o['geo_scope'],
            'country': o['country'] or '',
            'state': o['state'] or '',
            'state_code': o['state_code'] or '',
            'title': o['title'],
            'tech_category': o['tech_category'],
            'status': o['status'],
            'effective_date': o['effective_date'] or '',
            'risk_score': o['risk']['score'],
            'risk_rag': o['risk']['rag'],
            'criticality': o['criticality'],
            'evidence_status': o['evidence_status'],
            'source_count': len(o.get('provenance', [])),
        })

print(f'\nJSON report: {OUT_JSON}  ({OUT_JSON.stat().st_size // 1024} KB)')
print(f'Raw rows:    {OUT_RAW}')
print(f'Clean JSON:  {OUT_CLEAN_JSON}')
print(f'Clean CSV:   {OUT_CLEAN_CSV}')
print(f'\nStats:')
for k, v in report['stats'].items():
    if k != 'tech_breakdown':
        print(f'  {k}: {v}')
print('  Tech breakdown:', dict(tech_counts.most_common(8)))
print('\nDone.')
