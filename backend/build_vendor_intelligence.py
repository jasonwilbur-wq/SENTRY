"""Build SENTRY's canonical Vendor Intelligence directory from local sources.

Inputs:
- Vendor Assessments/Original Emerging Tech Trackers/*.xlsx
- Vendor Assessments/Vendor Assessment Reports/**

Outputs under Vendor Assessments/00_System:
- sentry_vendor_directory.csv              (app-ready canonical directory rows)
- sentry_vendor_report_inventory.csv       (all report files with selection flags)
- sentry_vendor_report_archive_manifest.csv(dry-run move plan for superseded docs)
- sentry_category_taxonomy.csv             (the <20 category taxonomy)

This script is non-destructive. It creates manifests; it does not delete or move
source documents. Archive moves can be applied later from the manifest after a
human spot-check.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from path_config import (
    VENDOR_ASSESSMENTS_ROOT,
    VENDOR_REPORTS_ROOT,
    VENDOR_SYSTEM_ROOT,
    VENDOR_TRACKERS_ROOT,
)
from vendor_taxonomy import CANONICAL_CATEGORIES, canonical_category, normalize_vendor_key, risk_from_rating

try:
    from docx import Document
except Exception:  # pragma: no cover - python-docx is optional for metadata-light builds
    Document = None

TRACKER_COLUMN_ALIASES = {
    "company": ("Company", "Vendor", "Company Name", "Cmpany"),
    "product": ("Technology_Product", "Technology Product", "Product", "Solution", "Offering"),
    "category": ("Category", "Technology Category"),
    "date": ("Date", "Assessment Date", "Last Assessed"),
    "status": ("Status", "Score"),
    "use_case": ("Use Case", "Use Cases"),
    "value": ("Add Value to Walmart", "Value to Walmart", "Business Value"),
    "maturity": ("Maturity Level", "Maturity"),
    "source_url": ("Source URL/Publisher", "Source URL", "Publisher"),
    "analysis_completed": ("Analysis Completed", "Complete Initial Assessment", "Initial Assessment"),
    "notes": ("Additional Notes", "Additional Notes2", "Initial Assessment Results"),
}

OUTPUT_DIRECTORY = "sentry_vendor_directory.csv"
OUTPUT_REPORTS = "sentry_vendor_report_inventory.csv"
OUTPUT_ARCHIVE = "sentry_vendor_report_archive_manifest.csv"
OUTPUT_TAXONOMY = "sentry_category_taxonomy.csv"
OUTPUT_SUMMARY = "sentry_vendor_build_summary.json"


@dataclass
class TrackerRecord:
    vendor_key: str
    company: str
    product: str = ""
    raw_category: str = ""
    canonical_category: str = "Other / Watchlist"
    source_month: str = ""
    source_workbook: str = ""
    source_sheet: str = ""
    date: str = ""
    status: str = ""
    use_case: str = ""
    value: str = ""
    maturity: str = ""
    source_url: str = ""
    analysis_completed: str = ""
    notes: str = ""
    rating: float = 3.0


@dataclass
class ReportRecord:
    vendor_folder: str
    vendor_key: str
    relative_path: str
    filename: str
    extension: str
    size_bytes: int
    modified_utc: str
    report_date: str = ""
    doc_title: str = ""
    doc_subject: str = ""
    doc_created: str = ""
    doc_modified: str = ""
    category_hint: str = "Other / Watchlist"
    is_var_like: bool = False
    is_draft_like: bool = False
    is_selected: bool = False
    base_vendor_key: str = ""
    base_vendor_name: str = ""
    service_label: str = ""


@dataclass
class DirectoryRow:
    vendor_key: str
    vendor_name: str
    product_key: str
    product_name: str
    canonical_category: str
    raw_tracker_category: str = ""
    source_evidence: str = "report"
    latest_tracker_month: str = ""
    tracker_status: str = ""
    rating: float = 3.0
    risk_level: str = "Medium"
    vendor_status: str = "Assessed"
    last_assessed: str = ""
    use_cases: str = ""
    value_to_walmart: str = ""
    maturity_level: str = ""
    source_url: str = ""
    notes: str = ""
    report_count: int = 0
    var_like_report_count: int = 0
    selected_report_path: str = ""
    selected_report_filename: str = ""
    selected_report_date: str = ""
    selected_report_modified_utc: str = ""
    archive_candidate_count: int = 0
    has_multiple_reports: bool = False
    source_workbooks: str = ""
    source_report_folders: str = ""
    confidence: str = "medium"
    needs_review: str = ""

    def to_csv_row(self, run_label: str, run_timestamp_utc: str, actor_id: str) -> dict[str, Any]:
        return {
            "run_label": run_label,
            "run_timestamp_utc": run_timestamp_utc,
            "actor_id": actor_id,
            "vendor_key": self.vendor_key,
            "vendor_name": self.vendor_name,
            "product_key": self.product_key,
            "product_name": self.product_name,
            "canonical_category": self.canonical_category,
            "raw_tracker_category": self.raw_tracker_category,
            "source_evidence": self.source_evidence,
            "latest_tracker_month": self.latest_tracker_month,
            "tracker_status": self.tracker_status,
            "rating": f"{self.rating:.2f}",
            "risk_level": self.risk_level,
            "vendor_status": self.vendor_status,
            "last_assessed": self.last_assessed,
            "use_cases": self.use_cases,
            "value_to_walmart": self.value_to_walmart,
            "maturity_level": self.maturity_level,
            "source_url": self.source_url,
            "notes": self.notes,
            "report_count": self.report_count,
            "var_like_report_count": self.var_like_report_count,
            "selected_report_path": self.selected_report_path,
            "selected_report_filename": self.selected_report_filename,
            "selected_report_date": self.selected_report_date,
            "selected_report_modified_utc": self.selected_report_modified_utc,
            "archive_candidate_count": self.archive_candidate_count,
            "has_multiple_reports": "YES" if self.has_multiple_reports else "",
            "source_workbooks": self.source_workbooks,
            "source_report_folders": self.source_report_folders,
            "confidence": self.confidence,
            "needs_review": self.needs_review,
        }


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def parse_float(value: str) -> float | None:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def month_from_filename(path: Path) -> str:
    match = re.search(r"(20\d{4})", path.name)
    return match.group(1) if match else ""


def date_from_filename(name: str) -> str:
    match = re.search(r"(20\d{2})(\d{2})(\d{2})", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    match = re.search(r"(20\d{2})[-_ ]?(\d{2})[-_ ]?(\d{2})", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return ""


def coalesce(row: dict[str, str], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        value = clean_text(row.get(alias))
        if value:
            return value
    return ""


def find_header_row(ws) -> tuple[int, list[str]]:
    expected = {alias.lower() for aliases in TRACKER_COLUMN_ALIASES.values() for alias in aliases}
    best: tuple[int, list[str], int] = (1, [], -1)
    for row_idx, raw_row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), start=1):
        values = [clean_text(v) for v in raw_row]
        score = sum(10 for v in values if v.lower() in expected) + sum(1 for v in values if v)
        if score > best[2]:
            while values and not values[-1]:
                values.pop()
            best = (row_idx, values, score)
    return best[0], [header or f"Column_{idx + 1}" for idx, header in enumerate(best[1])]


def tracker_rating(row: dict[str, str]) -> float:
    numeric = parse_float(coalesce(row, ("Status", "Score", "Overall Rating")))
    if numeric is not None:
        return max(0.0, min(5.0, numeric))

    status = coalesce(row, ("Status", "Score", "Analysis Completed", "Complete Initial Assessment")).lower()
    maturity = coalesce(row, ("Maturity Level", "Maturity")).lower()
    base = 3.0
    if "research further" in status or "reassess" in status:
        base = 3.2
    elif "reject" in status or "fail" in status:
        base = 2.0
    elif "pass" in status or "yes" in status or "complete" in status:
        base = 3.6
    elif "new" in status:
        base = 3.0

    if "mature" in maturity:
        base += 0.3
    elif "growth" in maturity:
        base += 0.1
    elif "early" in maturity or "concept" in maturity:
        base -= 0.2
    return round(max(0.0, min(5.0, base)), 2)


def read_trackers(trackers_root: Path) -> list[TrackerRecord]:
    records: list[TrackerRecord] = []
    for workbook in sorted(trackers_root.glob("*.xlsx")):
        wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row, headers = find_header_row(ws)
            if not headers:
                continue
            if not any(h in headers for h in ("Company", "Cmpany", "Vendor", "Company Name")):
                continue
            for raw_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
                values = list(raw_values[: len(headers)])
                row = {headers[idx]: clean_text(values[idx]) if idx < len(values) else "" for idx in range(len(headers))}
                company = coalesce(row, TRACKER_COLUMN_ALIASES["company"])
                product = coalesce(row, TRACKER_COLUMN_ALIASES["product"])
                raw_category = coalesce(row, TRACKER_COLUMN_ALIASES["category"])
                if not company or company.lower() in {"company", "vendor"}:
                    continue
                if not product and not raw_category:
                    continue

                category = canonical_category(raw_category, product, company, coalesce(row, TRACKER_COLUMN_ALIASES["use_case"]))
                records.append(TrackerRecord(
                    vendor_key=normalize_vendor_key(company),
                    company=company,
                    product=product,
                    raw_category=raw_category,
                    canonical_category=category,
                    source_month=month_from_filename(workbook),
                    source_workbook=workbook.name,
                    source_sheet=ws.title,
                    date=coalesce(row, TRACKER_COLUMN_ALIASES["date"]),
                    status=coalesce(row, TRACKER_COLUMN_ALIASES["status"]),
                    use_case=coalesce(row, TRACKER_COLUMN_ALIASES["use_case"]),
                    value=coalesce(row, TRACKER_COLUMN_ALIASES["value"]),
                    maturity=coalesce(row, TRACKER_COLUMN_ALIASES["maturity"]),
                    source_url=coalesce(row, TRACKER_COLUMN_ALIASES["source_url"]),
                    analysis_completed=coalesce(row, TRACKER_COLUMN_ALIASES["analysis_completed"]),
                    notes=coalesce(row, TRACKER_COLUMN_ALIASES["notes"]),
                    rating=tracker_rating(row),
                ))
    return records


def docx_core_props(path: Path) -> dict[str, str]:
    if path.suffix.lower() != ".docx":
        return {}
    try:
        with zipfile.ZipFile(path) as zf:
            raw = zf.read("docProps/core.xml").decode("utf-8", errors="ignore")
    except Exception:
        return {}
    props: dict[str, str] = {}
    for tag in ("title", "subject", "creator", "keywords", "description", "lastModifiedBy", "created", "modified"):
        match = re.search(rf"<(?:dc:|cp:|dcterms:)?{tag}[^>]*>(.*?)</(?:dc:|cp:|dcterms:)?{tag}>", raw, flags=re.S)
        if match:
            props[tag] = re.sub(r"\s+", " ", match.group(1)).strip()
    return props


def docx_excerpt(path: Path, max_chars: int = 1800) -> str:
    if Document is None or path.suffix.lower() != ".docx":
        return ""
    try:
        doc = Document(str(path))
        chunks: list[str] = []
        for paragraph in doc.paragraphs[:60]:
            text = paragraph.text.strip()
            if text:
                chunks.append(text)
            if sum(len(chunk) for chunk in chunks) >= max_chars:
                break
        return " ".join(chunks)[:max_chars]
    except Exception:
        return ""


def report_quality(record: ReportRecord) -> tuple[int, str, str, str]:
    """Higher tuple wins for selecting the active report per service."""
    path_lower = record.relative_path.lower()
    score = 0
    if record.is_var_like:
        score += 100
    if "vendor assessment report" in path_lower or "wmt-sec-var" in record.filename.lower():
        score += 30
    if record.extension == ".docx":
        score += 10
    if record.is_draft_like:
        score -= 25
    return (score, record.report_date, record.modified_utc, record.filename)


def service_label_from_folder(folder: str, base_name: str) -> str:
    label = folder.strip()
    if base_name and normalize_vendor_key(label).startswith(normalize_vendor_key(base_name)):
        # Preserve useful suffixes such as "Biostar X" or "Threatx SOC".
        suffix = label[len(base_name):].strip(" -_/.")
        if suffix:
            return suffix
    return "Primary Offering"


def read_reports(reports_root: Path) -> list[ReportRecord]:
    records: list[ReportRecord] = []
    for file in reports_root.rglob("*"):
        if not file.is_file():
            continue
        relative = file.relative_to(reports_root)
        if not relative.parts or relative.parts[0] == "00_Review_Archive":
            continue
        props = docx_core_props(file)
        excerpt = docx_excerpt(file)
        candidate = " ".join([file.name, str(relative), props.get("title", ""), props.get("subject", ""), excerpt])
        stat = file.stat()
        records.append(ReportRecord(
            vendor_folder=relative.parts[0],
            vendor_key=normalize_vendor_key(relative.parts[0]),
            relative_path=str(relative),
            filename=file.name,
            extension=file.suffix.lower(),
            size_bytes=stat.st_size,
            modified_utc=datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(timespec="seconds"),
            report_date=date_from_filename(file.name),
            doc_title=props.get("title", ""),
            doc_subject=props.get("subject", ""),
            doc_created=props.get("created", ""),
            doc_modified=props.get("modified", ""),
            category_hint=canonical_category("", candidate),
            is_var_like=bool(re.search(r"\bVAR\b|Vendor\s+Assessment\s+Report|WMT-SEC-VAR|Technical\s+Assessment", candidate, flags=re.I)),
            is_draft_like=bool(re.search(r"\bdraft\b", file.name, flags=re.I)),
        ))
    return records


def latest_trackers_by_vendor(records: list[TrackerRecord]) -> dict[str, TrackerRecord]:
    latest: dict[str, TrackerRecord] = {}
    for record in records:
        current = latest.get(record.vendor_key)
        if current is None or (record.source_month, record.date, record.source_workbook) >= (current.source_month, current.date, current.source_workbook):
            latest[record.vendor_key] = record
    return latest


def assign_report_base_vendors(reports: list[ReportRecord], trackers_by_key: dict[str, TrackerRecord]) -> None:
    tracker_keys = sorted(trackers_by_key, key=len, reverse=True)
    for report in reports:
        if report.vendor_key in trackers_by_key:
            base_key = report.vendor_key
        else:
            base_key = ""
            for candidate in tracker_keys:
                if len(candidate) >= 4 and report.vendor_key.startswith(candidate):
                    base_key = candidate
                    break
        if base_key:
            base = trackers_by_key[base_key]
            report.base_vendor_key = base.vendor_key
            report.base_vendor_name = base.company
            report.service_label = service_label_from_folder(report.vendor_folder, base.company)
        else:
            report.base_vendor_key = report.vendor_key
            report.base_vendor_name = report.vendor_folder
            report.service_label = "Primary Offering"


def product_key(vendor_key: str, product: str, report_folder: str = "") -> str:
    raw = f"{vendor_key}::{product or report_folder or 'primary'}"
    return hashlib.sha256(raw.lower().encode("utf-8")).hexdigest()[:12]


def combine_text(values: list[str], limit: int = 800) -> str:
    seen: list[str] = []
    for value in values:
        value = re.sub(r"\s+", " ", value or "").strip()
        if value and value not in seen:
            seen.append(value)
    return " | ".join(seen)[:limit]


def build_directory(trackers: list[TrackerRecord], reports: list[ReportRecord]) -> tuple[list[DirectoryRow], list[ReportRecord], list[dict[str, str]]]:
    latest_by_vendor = latest_trackers_by_vendor(trackers)
    assign_report_base_vendors(reports, latest_by_vendor)

    reports_by_base: dict[str, list[ReportRecord]] = defaultdict(list)
    for report in reports:
        reports_by_base[report.base_vendor_key].append(report)

    selected_by_folder: dict[str, ReportRecord] = {}
    archive_manifest: list[dict[str, str]] = []

    folder_groups: dict[str, list[ReportRecord]] = defaultdict(list)
    for report in reports:
        folder_groups[report.vendor_folder].append(report)

    for folder, folder_reports in folder_groups.items():
        candidates = [r for r in folder_reports if r.is_var_like] or folder_reports
        selected = max(candidates, key=report_quality)
        selected.is_selected = True
        selected_by_folder[folder] = selected
        for report in folder_reports:
            if report is selected:
                continue
            destination = Path("00_Review_Archive") / "Superseded_For_App" / report.vendor_folder / report.filename
            archive_manifest.append({
                "action": "move_to_review_archive",
                "reason": "Not selected for active Vendor Directory; newer or higher-quality report retained.",
                "vendor_folder": report.vendor_folder,
                "source_relative_path": report.relative_path,
                "destination_relative_path": str(destination),
                "selected_relative_path": selected.relative_path,
            })

    tracker_products: dict[tuple[str, str], list[TrackerRecord]] = defaultdict(list)
    for tracker in trackers:
        tracker_products[(tracker.vendor_key, tracker.product or "Primary Offering")].append(tracker)

    rows: list[DirectoryRow] = []
    emitted_products: set[tuple[str, str]] = set()
    emitted_product_keys: dict[str, set[str]] = defaultdict(set)

    for (vendor_key, product_name), product_trackers in sorted(tracker_products.items(), key=lambda item: (item[0][0], item[0][1])):
        latest = max(product_trackers, key=lambda r: (r.source_month, r.date, r.source_workbook))
        vendor_reports = reports_by_base.get(vendor_key, [])
        selected_reports = [r for r in vendor_reports if r.is_selected]
        category_votes = Counter({latest.canonical_category: 4})
        for report in selected_reports:
            category_votes[report.category_hint] += 1
        category = category_votes.most_common(1)[0][0]
        selected = max(selected_reports, key=report_quality) if selected_reports else None
        report_count = len(vendor_reports)
        var_count = sum(1 for r in vendor_reports if r.is_var_like)
        rows.append(DirectoryRow(
            vendor_key=vendor_key,
            vendor_name=latest.company,
            product_key=product_key(vendor_key, product_name),
            product_name=product_name,
            canonical_category=category,
            raw_tracker_category=latest.raw_category,
            source_evidence="tracker+report" if vendor_reports else "tracker",
            latest_tracker_month=latest.source_month,
            tracker_status=latest.status,
            rating=latest.rating,
            risk_level=risk_from_rating(latest.rating),
            vendor_status="Assessed" if selected else ("Under Review" if latest.analysis_completed.lower() != "yes" else "Screened"),
            last_assessed=latest.date or latest.source_month,
            use_cases=combine_text([r.use_case for r in product_trackers], 1200),
            value_to_walmart=combine_text([r.value for r in product_trackers], 1200),
            maturity_level=latest.maturity,
            source_url=latest.source_url,
            notes=combine_text([r.notes for r in product_trackers], 1200),
            report_count=report_count,
            var_like_report_count=var_count,
            selected_report_path=str(VENDOR_REPORTS_ROOT / selected.relative_path) if selected else "",
            selected_report_filename=selected.filename if selected else "",
            selected_report_date=selected.report_date if selected else "",
            selected_report_modified_utc=selected.modified_utc if selected else "",
            archive_candidate_count=sum(1 for r in vendor_reports if not r.is_selected),
            has_multiple_reports=report_count > 1,
            source_workbooks=combine_text(sorted({r.source_workbook for r in product_trackers}), 500),
            source_report_folders=combine_text(sorted({r.vendor_folder for r in vendor_reports}), 800),
            confidence="high" if selected and latest.raw_category else "medium",
            needs_review="" if category != "Other / Watchlist" else "category_review",
        ))
        emitted_products.add((vendor_key, product_name))
        emitted_product_keys[vendor_key].add(normalize_vendor_key(product_name))

    for folder, selected in sorted(selected_by_folder.items()):
        vendor_key = selected.base_vendor_key
        if vendor_key in latest_by_vendor:
            # A tracker row already represents the base vendor; add only service-specific selected reports
            # when the folder has a meaningful offering suffix not covered by a tracker product.
            service = selected.service_label or "Primary Offering"
            service_key = normalize_vendor_key(service)
            existing_product_keys = emitted_product_keys.get(vendor_key, set())
            if service == "Primary Offering" or (vendor_key, service) in emitted_products:
                continue
            if any(service_key and (service_key in existing or existing in service_key) for existing in existing_product_keys):
                continue
            base = latest_by_vendor[vendor_key]
            vendor_name = base.company
            raw_category = base.raw_category
            source_month = base.source_month
            rating = base.rating
            use_case = base.use_case
            value = base.value
            maturity = base.maturity
            source_url = base.source_url
            notes = base.notes
        else:
            vendor_name = selected.base_vendor_name
            service = selected.service_label or "Primary Offering"
            raw_category = ""
            source_month = ""
            rating = 3.2 if selected.is_var_like else 3.0
            use_case = ""
            value = ""
            maturity = ""
            source_url = ""
            notes = selected.doc_subject or selected.doc_title

        vendor_reports = reports_by_base.get(vendor_key, [])
        folder_reports = folder_groups.get(folder, [])
        category = selected.category_hint
        rows.append(DirectoryRow(
            vendor_key=vendor_key,
            vendor_name=vendor_name,
            product_key=product_key(vendor_key, service, folder),
            product_name=service,
            canonical_category=category,
            raw_tracker_category=raw_category,
            source_evidence="tracker+report" if source_month else "report",
            latest_tracker_month=source_month,
            rating=rating,
            risk_level=risk_from_rating(rating),
            vendor_status="Assessed" if selected.is_var_like else "Report Linked",
            last_assessed=selected.report_date or selected.modified_utc[:10],
            use_cases=use_case,
            value_to_walmart=value,
            maturity_level=maturity,
            source_url=source_url,
            notes=notes,
            report_count=len(folder_reports),
            var_like_report_count=sum(1 for r in folder_reports if r.is_var_like),
            selected_report_path=str(VENDOR_REPORTS_ROOT / selected.relative_path),
            selected_report_filename=selected.filename,
            selected_report_date=selected.report_date,
            selected_report_modified_utc=selected.modified_utc,
            archive_candidate_count=sum(1 for r in folder_reports if not r.is_selected),
            has_multiple_reports=len(folder_reports) > 1,
            source_workbooks="",
            source_report_folders=folder,
            confidence="medium" if category != "Other / Watchlist" else "low",
            needs_review="" if category != "Other / Watchlist" else "category_review",
        ))

    rows.sort(key=lambda r: (r.vendor_name.lower(), r.product_name.lower()))
    return rows, reports, archive_manifest


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_outputs(rows: list[DirectoryRow], reports: list[ReportRecord], archive_manifest: list[dict[str, str]], actor_id: str) -> dict[str, Any]:
    VENDOR_SYSTEM_ROOT.mkdir(parents=True, exist_ok=True)
    run_label = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    directory_rows = [row.to_csv_row(run_label, run_timestamp, actor_id) for row in rows]
    directory_fields = list(directory_rows[0].keys()) if directory_rows else []
    write_csv(VENDOR_SYSTEM_ROOT / OUTPUT_DIRECTORY, directory_rows, directory_fields)

    report_rows = [{
        "vendor_folder": r.vendor_folder,
        "base_vendor_key": r.base_vendor_key,
        "base_vendor_name": r.base_vendor_name,
        "service_label": r.service_label,
        "relative_path": r.relative_path,
        "filename": r.filename,
        "extension": r.extension,
        "size_bytes": r.size_bytes,
        "modified_utc": r.modified_utc,
        "report_date": r.report_date,
        "category_hint": r.category_hint,
        "is_var_like": "YES" if r.is_var_like else "",
        "is_draft_like": "YES" if r.is_draft_like else "",
        "is_selected_for_directory": "YES" if r.is_selected else "",
        "doc_title": r.doc_title,
        "doc_subject": r.doc_subject,
        "doc_created": r.doc_created,
        "doc_modified": r.doc_modified,
    } for r in reports]
    write_csv(VENDOR_SYSTEM_ROOT / OUTPUT_REPORTS, report_rows, list(report_rows[0].keys()) if report_rows else [])

    archive_fields = ["action", "reason", "vendor_folder", "source_relative_path", "destination_relative_path", "selected_relative_path"]
    write_csv(VENDOR_SYSTEM_ROOT / OUTPUT_ARCHIVE, archive_manifest, archive_fields)

    taxonomy_rows = [{"sort_order": idx + 1, "category": category} for idx, category in enumerate(CANONICAL_CATEGORIES)]
    write_csv(VENDOR_SYSTEM_ROOT / OUTPUT_TAXONOMY, taxonomy_rows, ["sort_order", "category"])

    summary = {
        "run_label": run_label,
        "run_timestamp_utc": run_timestamp,
        "actor_id": actor_id,
        "source_roots": {
            "vendor_assessments_root": str(VENDOR_ASSESSMENTS_ROOT),
            "trackers_root": str(VENDOR_TRACKERS_ROOT),
            "reports_root": str(VENDOR_REPORTS_ROOT),
            "system_root": str(VENDOR_SYSTEM_ROOT),
        },
        "directory_rows": len(rows),
        "logical_vendors": len({row.vendor_key for row in rows}),
        "selected_reports": sum(1 for report in reports if report.is_selected),
        "report_files_indexed": len(reports),
        "archive_candidates": len(archive_manifest),
        "category_count": len({row.canonical_category for row in rows}),
        "categories": Counter(row.canonical_category for row in rows).most_common(),
        "evidence_mix": Counter(row.source_evidence for row in rows).most_common(),
        "needs_review": Counter(row.needs_review or "none" for row in rows).most_common(),
    }
    (VENDOR_SYSTEM_ROOT / OUTPUT_SUMMARY).write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build SENTRY canonical Vendor Intelligence files")
    parser.add_argument("--actor-id", default="code-master-129743", help="Actor ID written into generated CSV metadata")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    trackers = read_trackers(VENDOR_TRACKERS_ROOT)
    reports = read_reports(VENDOR_REPORTS_ROOT)
    rows, reports, archive_manifest = build_directory(trackers, reports)
    summary = write_outputs(rows, reports, archive_manifest, actor_id=args.actor_id)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
